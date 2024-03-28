#!/usr/bin/env python
# coding: utf-8

# In[2]:


# no need to change 
# IF WARNING, RERUN CUBE AGAIN

#import urllib.request
import numpy as np, pandas as pd, difflib
from pandas import DataFrame
import re

import fuzzywuzzy 
from fuzzywuzzy import fuzz
from fuzzywuzzy import process

import openpyxl
from openpyxl.descriptors import (String,Sequence,Integer)
from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.styles import numbers,Alignment,PatternFill,Font

def same(s1, s2):
    return difflib.SequenceMatcher(lambda x:x==(' ','x','X'), s1, s2).quick_ratio()

from gensim.models import FastText
from fractions import Fraction
from sklearn.svm import SVC
from sklearn.model_selection import RandomizedSearchCV
from sklearn.metrics import classification_report
from sklearn.model_selection import train_test_split
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.model_selection import ShuffleSplit

# import time for efficiency
import time
import calendar
import datetime


# In[159]:


#pd.set_option('display.max_rows',None)


# In[3]:


# help format the list in new dic 
def add_to_list(dictionary, key, new_element):
    
    if key in dictionary:
        dictionary[key].append(new_element) 
    else:
        dictionary[key] = [new_element]


# 2个问题：
# 1.按名字修改，直接给所有相同名字的都改了
# 2.alternative的sku什么的，怎么办
# 

# In[4]:


# Enter which month we want to report, E.G. Jan for January
month=input('Which Month to report: (like Jan for January) ').strip().capitalize()

# Enter which COGS file we want to use as data source, using file path in computer
cogs=input('What is the COGS File path: (like C:\FI\COGS - December 2022.xlsx) ').strip()

# Enter the sheet name of master list we using 
mst_sheet=input('What is the sheet name for last month in master list: (like Jan) ').strip().capitalize()


# In[5]:



# Enter the target data resource for master list
mst_path=input("What is the Master List File path: (like \\VRCCO-24\\Users\Michael\Desktop\INV EXCEL FILES\Monthly Inv Report.xlsx) ").strip() or r"\\VRCCO-24\Users\Michael\Desktop\INV EXCEL FILES\Monthly Inv Report.xlsx" 


# In[6]:


# Enter the file path of purchase history
# The sheet order should be MWI>MK>MILA>EQ>ABR>EST
purchase=input('What is the file path of purchase history: (like C:\Inv Data\Purchase History\Purchase History.xlsx) ').strip() or r'C:\Inv Data\Purchase History\Purchase History.xlsx'


# In[7]:


if month=='Jan':
    mon='01'
elif month=='Feb':
    mon='02'
elif month=='Mar':
    mon='03'
elif month=='Apr':
    mon='04'
elif month=='May':
    mon='05'
elif month=='Jun':
    mon='06'
elif month=='Jul':
    mon='07'
elif month=='Aug':
    mon='08'
elif month=='Sep':
    mon='09'
elif month=='Oct':
    mon='10'
elif month=='Nov':
    mon='11'
elif month=='Dec':
    mon='12'


# In[8]:


# Input the data base here with the location, clean up the data first. Delete the useless header
# Input the renew data
renew=pd.read_excel('C:\\Inv Data\\Monthly Renew.xlsx',converters={'sku':str,'new sku':str})

# Input the 'master list' template with the basic info
mst_list=pd.read_excel(mst_path,sheet_name=mst_sheet,converters={'SKU':str})
mst_list['Product']=mst_list['Product'].str.rstrip()

last=mst_list.copy()

# Open purchase amount file
# ABR MILA ES don't provide purchase hist, might delete code
mwi_qty=pd.read_excel(purchase,sheet_name=0,converters={'MWI SKU':str})
mwi_qty=mwi_qty.dropna(subset=['Description'])
mwi_qty=mwi_qty.drop( mwi_qty[mwi_qty['Qty'] == 0].index.tolist())
mwi_qty=mwi_qty.reset_index(drop=True)
mk_qty=pd.read_excel(purchase,sheet_name=1,converters={'Item #':str,'Price':str})
mk_qty=mk_qty.dropna(subset=['Description'])
mk_qty=mk_qty.reset_index(drop=True)
mila_qty=pd.read_excel(purchase,sheet_name=2,converters={'sku':str})
eq_qty=pd.read_excel(purchase,sheet_name=3)
abr_qty=pd.read_excel(purchase,sheet_name=4)
es_qty=pd.read_excel(purchase,sheet_name=5,converters={'sku':str})

# Input Invoice Cost Report from Bre
bre=pd.read_excel(cogs)

bre=bre.drop( bre[bre['Products'] == 0].index.tolist() )
#except IndexError:
#    bre=bre
bre=bre.dropna(subset=['Products'])
# Delete the data for other departments based on suppliers
v1=bre.loc[bre['Vendors']=='Amatheon Animal Health'].index.tolist()
v2=bre.loc[bre['Vendors']=='Boehringer Ingelheim Animal Health USA Inc.'].index.tolist()
v3=bre.loc[bre['Vendors']=='Costco'].index.tolist()
v4=bre.loc[bre['Vendors']=='eBay'].index.tolist()
v5=bre.loc[bre['Vendors']=='Stokes Healthcare'].index.tolist()
v6=bre.loc[bre['Vendors']=='Greer Laboratories, Inc'].index.tolist()
v7=bre.loc[bre['Vendors']=='Heska Corporation'].index.tolist()
v8=bre.loc[bre['Vendors']=='Hills Pet Nutrition Inc'].index.tolist()
v9=bre.loc[bre['Vendors']=='Norco Inc'].index.tolist()
v10=bre.loc[bre['Vendors']=='Royal Canin'].index.tolist()
v11=bre.loc[bre['Vendors']=='Veterinary Orthopedic Implants'].index.tolist()
v12=bre.loc[bre['Vendors']=='Wedgewood'].index.tolist()
v13=bre.loc[bre['Vendors']=='Veterinary Solutions Direct'].index.tolist()
v14=bre.loc[bre['Vendors']=='Stokes Healthcare'].index.tolist()
v15=bre.loc[bre['Vendors']=='Biomedtrix'].index.tolist()
v16=bre.loc[bre['Vendors']=='0'].index.tolist()
v17=bre.loc[bre['Vendors']=='Elanco'].index.tolist()
v18=bre.loc[bre['Vendors']=='Zoetis'].index.tolist()
v19=bre.loc[bre['Vendors']=='MSU'].index.tolist()
v20=bre.loc[bre['Vendors']=='Torigen Pharmaceuticals'].index.tolist()
v0=bre[bre['Products'].str.contains('MG/')].index.tolist()+bre[bre['Products'].str.contains('mg/')].index.tolist()+bre[bre['Products'].str.contains('IU/')].index.tolist()+bre[bre['Products'].str.contains('TAB ')].index.tolist()
other_ven=v0+v1+v2+v3+v4+v5+v6+v7+v8+v9+v10+v11+v12+v13+v14+v15+v16+v17+v18+v19+v20
bre.drop(other_ven,inplace=True)
bre=bre.reset_index(drop=True)

pd.set_option('display.width', 50)


# In[9]:


# Set up link for purchase history and COGS. We record one product for both report

# create list for same product
one={}

# merge the purchase history, some supplier don't provide purchase history, so we just ignore them while comparing similarity
# find and eliminate pharmacy items for better performance
# name: qty, price
hist={}
pha={'MWI':[],'MK':[]}
for h1 in enumerate(mwi_qty['Description']):
    if h1[1].__contains__('Tablet') or h1[1].__contains__('Caplet') or h1[1].__contains__('mg/') or h1[1].__contains__('IU/') or h1[1].__contains__('Capsule') or h1[1].__contains__('gm/') or h1[1].__contains__('mcg'):
        pha['MWI'].append(h1[0])
        continue
    else:
        hist[h1[1]]=[[mwi_qty['Qty'][h1[0]]],[mwi_qty['Current Price'][h1[0]]]]              
for h2 in enumerate(mk_qty['Description']):
    if h2[1].__contains__('MG'):
        pha['MK'].append(h2[0])
        continue
    else:
        hist[h2[1]]=[[mk_qty[mk_qty['Description']==h2[1]]['PO Qty'].sum()],[float(mk_qty[mk_qty['Description']==h2[1]].iloc[-1]['Price'].replace(',',''))]]
for h3 in enumerate(eq_qty['Product Desc']):
    if h3[1] in hist:
        hist[h3[1]][0]=[hist[h3[1]][0][0]+1]
        hist[h3[1]][1]=[eq_qty['Total'][h3[0]]]
    else:
        hist[h3[1]]=[[1],[eq_qty['Total'][h3[0]]]]
for h4 in enumerate(mila_qty['name']):
    hist[h4[1]]=[[mila_qty['qty'][h4[0]]],[mila_qty['price'][h4[0]]]]
for h5 in enumerate(es_qty['name']):
    hist[h5[1]]=[[es_qty['qty'][h5[0]]],[es_qty['price'][h5[0]]]]
        
# combinate 2 dict, just join hist + es
#hist=hist|es
# start to compare with COGS

for item1 in hist.keys():
    sc=0
    for item2 in enumerate(bre['Products']):
        vendor=bre['Vendors'][item2[0]]
        if vendor=='ESutures' or vendor=='McKesson' or vendor=='MWI'or vendor=='Equasheild LLC' or vendor=='MILA International, Inc':
            nsc=0.8*fuzz.token_set_ratio(str(item2[1]),str(item1))+0.2*fuzz.partial_ratio(str(item2[1]),str(item1))
            qty_hist=float(hist[item1][0][0])
            qty_bre=float(bre['Quantity'][item2[0]])
            p_hist=float(hist[item1][1][0])
            p_bre=float(bre['Price'][item2[0]])

            # give more power for the prod with same qty and price, which means they are more likely the same product   
            if qty_hist==qty_bre:
                nsc+=4
            if p_hist==p_bre:
                nsc+=6
            if qty_hist==qty_bre and p_hist==p_bre:
                nsc+=20

            # supplier : bre
            if nsc>=70 and nsc>sc:
                one[item1]=item2[1]
                sc=nsc


# In[10]:


# drop pharmacy item from purchase hist
mwi_qty.drop(pha['MWI'],inplace=True)
mwi_qty=mwi_qty.reset_index(drop=True)
mk_qty.drop(pha['MK'],inplace=True)
mk_qty=mk_qty.reset_index(drop=True)


# In[11]:


mwiname={'Description':{},'MWI SKU':{},'Unit':{},'Qty':{},'Current Price':{}}


# In[12]:


cc=0
for y in enumerate(mwi_qty['Description']):
    if y[1] not in mwiname['Description'].values():
        mwiname['Description'][cc]=y[1]
        mwiname['MWI SKU'][cc]=str(mwi_qty['MWI SKU'][y[0]])
        mwiname['Unit'][cc]=mwi_qty['Unit'][y[0]]
        mwiname['Qty'][cc]=int(mwi_qty['Qty'][y[0]])
        mwiname['Current Price'][cc]=mwi_qty['Current Price'][y[0]]
        cc+=1
    else:
        ind=list(mwiname['Description'].keys())[list(mwiname['Description'].values()).index(y[1])]
        mwiname['Qty'][ind]+=int(mwi_qty.loc[[y[0]],['Qty']]['Qty'].tolist()[0])

mwiname=DataFrame(mwiname)


# In[13]:


mwi_qty=mwiname


# In[14]:


# Renew the master list first, to make sure everything we want to find is there

for z in enumerate(renew['name (required)']):
    # add new
    if pd.isnull(renew['alt (Y/N)'][z[0]]) or renew['alt (Y/N)'][z[0]]=='N' or renew['alt (Y/N)'][z[0]]=='n': 
        if z[1] not in mst_list['Product'].values:
            mst_list.loc[len(mst_list.index),'Supplier']=renew.loc[z[0],'new supplier']
            mst_list.loc[len(mst_list.index)-1,'Product']=renew.loc[z[0],'name (required)']
            mst_list.loc[len(mst_list.index)-1,'SKU']=renew.loc[z[0],'new sku']
            mst_list.loc[len(mst_list.index)-1,'Bin Qty']=renew.loc[z[0],'new bin qty']
            mst_list.loc[len(mst_list.index)-1,'Bin Size']=renew.loc[z[0],'new bin size']
            mst_list.loc[len(mst_list.index)-1,'Unit Qty']=renew.loc[z[0],'new unit qty']
            mst_list.loc[len(mst_list.index)-1,'Unit']=renew.loc[z[0],'new unit']
            mst_list.loc[len(mst_list.index)-1,'Location']=renew.loc[z[0],'new location']
            mst_list.loc[len(mst_list.index)-1,'Group']=renew.loc[z[0],'new group']

        # update old    
        else:
            # use loc add row, use index to list to find the proper row index
            ind=mst_list[mst_list['Product']==z[1]].index.tolist()[0]
            if pd.notnull(renew['new name'][z[0]]):
                mst_list.loc[[ind],['Product']]=renew['new name'][z[0]]
            if pd.notnull(renew['new sku'][z[0]]):
                mst_list.loc[[ind],['SKU']]=renew['new sku'][z[0]]
            if pd.notnull(renew['new bin qty'][z[0]]):
                mst_list.loc[[ind],['Bin Qty']]=renew['new bin qty'][z[0]]    
            if pd.notnull(renew['new bin size'][z[0]]):
                mst_list.loc[[ind],['Bin Size']]=renew['new bin size'][z[0]] 
            if pd.notnull(renew['new unit'][z[0]]):
                mst_list.loc[[ind],['Unit']]=renew['new unit'][z[0]] 
            if pd.notnull(renew['new unit qty'][z[0]]):
                mst_list.loc[[ind],['Unit Qty']]=renew['new unit qty'][z[0]] 
            if pd.notnull(renew['new supplier'][z[0]]):
                mst_list.loc[[ind],['Supplier']]=renew['new supplier'][z[0]] 
            if pd.notnull(renew['new location'][z[0]]):
                mst_list.loc[[ind],['Location']]=renew['new location'][z[0]]  
            if pd.notnull(renew['new group'][z[0]]):
                mst_list.loc[[ind],['Group']]=renew['new group'][z[0]]  
                
    else:
        # put the alternative supplier info into the product
        ind=mst_list[mst_list['Product']==z[1]].index.tolist()[0]
        mst_list.loc[len(mst_list.index),'Product']=renew.loc[z[0],'name (required)']
        if pd.notnull(renew['new name'][z[0]]):
            mst_list.loc[len(mst_list.index)-1,'Product']=renew.loc[z[0],'new name']
        else:
            mst_list.loc[len(mst_list.index)-1,'Product']=mst_list.loc[ind,'Product']
        if pd.notnull(renew['new sku'][z[0]]):
            mst_list.loc[len(mst_list.index)-1,'SKU']=renew.loc[z[0],'new sku']
        else:
            mst_list.loc[len(mst_list.index)-1,'SKU']=mst_list.loc[ind,'SKU']
        if pd.notnull(renew['new bin qty'][z[0]]):
            mst_list.loc[len(mst_list.index)-1,'Bin Qty']=renew.loc[z[0],'new bin qty']
        else:
            mst_list.loc[len(mst_list.index)-1,'Bin Qty']=mst_list.loc[ind,'Bin Qty']
        if pd.notnull(renew['new bin size'][z[0]]):
            mst_list.loc[len(mst_list.index)-1,'Bin Size']=renew.loc[z[0],'new bin size']
        else:
            mst_list.loc[len(mst_list.index)-1,'Bin Size']=mst_list.loc[ind,'Bin Size']
        if pd.notnull(renew['new unit'][z[0]]):
            mst_list.loc[len(mst_list.index)-1,'Unit']=renew.loc[z[0],'new unit']
        else:
            mst_list.loc[len(mst_list.index)-1,'Unit']=mst_list.loc[ind,'Unit']
        if pd.notnull(renew['new unit qty'][z[0]]):
            mst_list.loc[len(mst_list.index)-1,'Unit Qty']=renew.loc[z[0],'new unit qty']
        else:
            mst_list.loc[len(mst_list.index)-1,'Unit Qty']=mst_list.loc[ind,'Unit Qty']
        if pd.notnull(renew['new supplier'][z[0]]):
            mst_list.loc[len(mst_list.index)-1,'Supplier']=renew.loc[z[0],'new supplier']
        else:
            mst_list.loc[len(mst_list.index)-1,'Supplier']=mst_list.loc[ind,'Supplier']
        mst_list.loc[len(mst_list.index)-1,'Location']=mst_list.loc[ind,'Location']
        mst_list.loc[len(mst_list.index)-1,'Group']=mst_list.loc[ind,'Group']
        mst_list.loc[len(mst_list.index)-1,'Alternative']=renew['alt (Y/N)'][z[0]]  
        
# poka-yoke
mst_list=mst_list.dropna(subset=['Supplier'])
mst_list=mst_list.reset_index(drop=True)


# In[15]:


# Different from drag data sheet, we just want the data that can be found in our report
# The dict shows the prod name and the SKU with the similarity score 
new={}
na=[]
used=[]

# Sort by suppliers. MWI>MK>others
# Start with the master list, use our prod as target

for i in enumerate(mst_list['Product']):    
    score=-1
    # SKU first, if cannot find SKU, we save for later check
    if mst_list['Supplier'][i[0]].replace(' ','')=='MWI':
        try:
            index=mwi_qty[mwi_qty['MWI SKU']==mst_list['SKU'][i[0]]].index[0]
            sku=mwi_qty['MWI SKU'][index]
            name=mwi_qty['Description'][index]
            price=mwi_qty['Current Price'][index]
            unit=mwi_qty['Unit'][index]
            qty=mwi_qty['Qty'][index]
            source='MWI Hist'
            used.append(name)
            add_to_list(new,i[1],[name,sku,price,unit,qty,100,source])

        except IndexError:
            na.append((i,'MWI'))
                
    elif mst_list['Supplier'][i[0]].replace(' ','')=='MK':
        try:
            index=mk_qty[mk_qty['Item #']==mst_list['SKU'][i[0]]].index[-1]
            sku=mk_qty['Item #'][index]
            name=mk_qty['Description'][index]
            price=float(mk_qty['Price'][index].replace(',',''))
            unit=mk_qty['UOM'][index]
            qty=mk_qty[mk_qty['Item #']==sku]['PO Qty'].sum()
            source='MK Hist'
            used.append(name)
            add_to_list(new,i[1],[name,sku,price,unit,qty,100,source])
        except IndexError:
            na.append((i,'MK'))
                
    elif mst_list['Supplier'][i[0]].replace(' ','')=='MILA':
        try:
            index=mila_qty[mila_qty['sku']==mst_list['SKU'][i[0]]].index[0]
            sku=mila_qty['sku'][index]
            name=mila_qty['name'][index]
            price=mila_qty['price'][index]
            unit=None
            qty=mila_qty['qty'][index]
            source='MILA Hist'
            used.append(name)
            add_to_list(new,i[1],[name,sku,price,unit,qty,100,source])
        except IndexError:
            #mst_list.loc[[i[0]],['Supplier Name']]=i[1]
            na.append((i,'MILA'))
                
    elif mst_list['Supplier'][i[0]].replace(' ','')=='Equashield':
        try:
            index=eq_qty[eq_qty['Product ID']==mst_list['SKU'][i[0]]].sort_values(by=['Invoice Date']).index[-1]
            sku=eq_qty['Product ID'][index]
            name=eq_qty['Product Desc'][index]
            price=eq_qty['Total'][index]
            unit='BX'
            qty=eq_qty[eq_qty['Product ID']==sku].count()[0]
            source='EQ Hist'
            used.append(name)
            add_to_list(new,i[1],[name,sku,price,unit,qty,100,source])
        except IndexError:
            na.append((i,'Equashield'))
                
                
    elif mst_list['Supplier'][i[0]].replace(' ','')=='E-Sutures':
        try:
            index=es_qty[es_qty['sku']==mst_list['SKU'][i[0]]].index[0]
            sku=es_qty['sku'][index]
            name=es_qty['name'][index]
            qty=es_qty['qty'][index]
            unit=es_qty['unit'][index]
            price=es_qty['price'][index]
            source='EST Hist'
            used.append(name)
            add_to_list(new,i[1],[name,sku,price,unit,qty,100,source])
        except IndexError:
            na.append((i,'E-Sutures'))
            
    else:
        na.append((i,mst_list['Supplier'][i[0]].replace(' ','')))
        


# In[16]:


# find the 100% 'same' prodcut, and delete it from COGS to avoid later misleading
drop_list=[]
for item3 in used:
    try:
        drop_list.append(bre[bre['Products']==one[item3]].index[0])
    except KeyError:
        continue
bre.drop(drop_list,inplace=True)
bre=bre.reset_index(drop=True)


# In[17]:


# search the product that cannot be found in purchase history, then search in COGS
used1=[]
remove=set()

mst_list['Supplier Name'].fillna(mst_list['Product'],inplace=True)
for n in na:
    product = mst_list['Supplier Name'][n[0][0]]
    score = 0
    if n[1] == 'MWI':
        source = 'MWI Hist'
        for j in enumerate(mwi_qty['Description']):
            if j[1] in used:
                continue
            else:
                new_score = 0.8 * fuzz.token_set_ratio(str(product), str(j[1])) + 0.2 * fuzz.partial_ratio(str(product),str(j[1]))
                if new_score > 65 and new_score > score:
                    try:
                        name = j[1]
                        price = mwi_qty['Current Price'][j[0]]
                        sku = mwi_qty['MWI SKU'][j[0]]
                        unit = mwi_qty['Unit'][j[0]]
                        qty = mwi_qty['Qty'][j[0]]
                        score = new_score
                        if score >99.5 :
                            used1.append(name)
                            add_to_list(new,n[0][1],[name, sku, price, unit, qty, score, source])
                            remove.add(n)
                            break
                    except IndexError:
                        continue
        if score != 100 and score != 0:
            add_to_list(new,n[0][1],[name, sku, price, unit, qty, score, source])
            remove.add(n)

    elif n[1] == 'MK':
        source = 'MK Hist'
        for k in enumerate(mk_qty['Description']):
            if k[1] in used:
                continue
            else:
                new_score = 0.8 * fuzz.token_set_ratio(product, str(k[1])) + 0.2 * fuzz.partial_ratio(product,str(k[1]))
                if new_score > 65 and new_score > score:
                    try:
                        name = k[1]
                        index = mk_qty[mk_qty['Description'] == name].index[-1]
                        price = float(mk_qty['Price'][index].replace(',', ''))
                        sku = mk_qty['Item #'][index]
                        unit = mk_qty['UOM'][index]
                        qty = mk_qty[mk_qty['Description'] == name]['PO Qty'].sum()
                        score = new_score
                        if score >99.5:
                            used1.append(name)
                            add_to_list(new,n[0][1],[name, sku, price, unit, qty, score, source])
                            remove.add(n)
                            break
                    except IndexError:
                        continue
        if score != 100 and score != 0:
            add_to_list(new,n[0][1],[name, sku, price, unit, qty, score, source])
            remove.add(n)

    elif n[1] == 'E-Sutures':
        source = 'EST Hist'
        for p in enumerate(es_qty['name']):
            if p[1] in used:
                continue
            else:
                new_score = 0.8 * fuzz.token_set_ratio(product, str(p[1])) + 0.2 * fuzz.partial_ratio(product,str(p[1]))
                if new_score > 65 and new_score > score:
                    try:
                        name = p[1]
                        price = es_qty['price'][p[0]]
                        sku = es_qty['sku'][p[0]]
                        unit = es_qty['unit'][p[0]]
                        qty = es_qty['qty'][p[0]]
                        score = new_score
                        if score >99.5:
                            used1.append(name)
                            add_to_list(new,n[0][1],[name, sku, price, unit, qty, score, source])
                            remove.add(n)
                            break
                    except IndexError:
                        continue
        if score != 100 and score != 0:
            add_to_list(new,n[0][1],[name, sku, price, unit, qty, score, source])
            remove.add(n)

    elif n[1] == 'MILA':
        source = 'MILA Hist'
        for q in enumerate(mila_qty['name']):
            if q[1] in used:
                continue
            else:
                new_score = 0.8 * fuzz.token_set_ratio(product, str(q[1])) + 0.2 * fuzz.partial_ratio(product, str(q[1]))
                if new_score > 65 and new_score > score:
                    try:
                        name = q[1]
                        price = mila_qty['price'][p[0]]
                        sku = mila_qty['sku'][p[0]]
                        unit = None
                        qty = mila_qty['qty'][p[0]]
                        score = new_score
                        if score >99.5:
                            used1.append(name)
                            add_to_list(new,n[0][1],[name, sku, price, unit, qty, score, source])
                            remove.add(n)
                            break
                    except IndexError:
                        continue
        if score != 100 and score != 0:
            add_to_list(new,n[0][1],[name, sku, price, unit, qty, score, source])
            remove.add(n)

    elif n[1] == 'ABR':
        source = 'ABR Hist'
        for m in enumerate(abr_qty['Memo']):
            new_score = 0.8 * fuzz.token_set_ratio(product, str(m[1])) + 0.2 * fuzz.partial_ratio(product, str(m[1]))
            if new_score > 65 and new_score > score:
                try:
                    name = m[1]
                    price = abr_qty['Sales Price'][m[0]]
                    sku = None
                    unit = abr_qty[abr_qty['Memo'] == name].iloc[-1]['U/M']
                    qty = abr_qty[abr_qty['Memo'] == name]['Qty'].sum()
                    score = new_score
                    if score >99.5:
                        used1.append(name)
                        add_to_list(new,n[0][1],[name, sku, price, unit, qty, score, source])
                        remove.add(n)
                        break
                except IndexError:
                    continue
        if score != 100 and score != 0:
            add_to_list(new,n[0][1],[name, sku, price, unit, qty, score, source])
            remove.add(n)
            
            


# In[18]:


# drop the used item again for better performance
drop_list=[]
for item4 in used1:
    try:
        drop_list.append(bre[bre['Products']==one[item4]].index[0])
    except KeyError:
        continue
bre.drop(drop_list,inplace=True)
bre=bre.reset_index(drop=True)

na_set=set(na)
na=list(na_set-remove)


# In[19]:


# search product by supplier name in COGS report
used2=[]
source='COGS'
for miss in na:    
    product=mst_list['Supplier Name'][miss[0][0]]
    supplier=miss[1]
    score=-1
    for o in enumerate(bre['Products']):
        # To avoid mistakes, like wrong input, we will rely on purchase history first, so ignore prod from the same supplier
        if o in used2:
            continue
        else:
            vendor=bre['Vendors'][o[0]]
            if vendor=='MWI' and supplier=='MWI':
                continue
            elif vendor=='MILA International, Inc' and supplier=='MILA':
                continue
            elif vendor=='McKesson' and supplier=='MK':
                continue
            elif vendor=='Equasheild LLC' and supplier=='Equashield':
                continue
            elif vendor=='ESutures' and supplier=='E-Sutures':
                continue 
            else:
                new_score=0.8*fuzz.token_set_ratio(product,str(o[1]))+0.2*fuzz.partial_ratio(product,str(o[1]))
                if new_score>score:
                    name=o[1]
                    price=bre['Price'][o[0]]
                    sku=mst_list['Supplier SKU'][miss[0][0]]
                    qty=bre['Quantity'][o[0]]
                    score=new_score
                    if score >99.5:
                        used2.append(o)
                        break
    if score>65:
        add_to_list(new,miss[0][1],[name,sku,price,mst_list['Pur Unit'][miss[0][0]],qty,score,source])
        
    else:  
        add_to_list(new,miss[0][1],[mst_list['Supplier Name'][miss[0][0]],mst_list['Supplier SKU'][miss[0][0]],mst_list['Price'][miss[0][0]],mst_list['Pur Unit'][miss[0][0]],float(0),None,None])


# In[20]:


# reset everything
mst_list['Pur Unit']=None
mst_list['Pur Qty']=None
mst_list['Bin Turn']=None
mst_list['Similarity']=None
mst_list['Source']=None
mst_list['Date']=None
mst_list['Growth R']=None

shown={}
# write down name/SKU/price...
for x in enumerate(mst_list['Product']):
    prod=x[1]
    id=x[0]
    
    # no other list inside
    if len(new[prod]) == 7:
        shown[prod]=0
        mst_list.loc[id,'Supplier SKU']=new[prod][1]
        mst_list.loc[id,'Supplier Name']=new[prod][0].rstrip()
        mst_list.loc[id,'Price']=new[prod][2]
        mst_list.loc[id,'Similarity']=new[prod][5]
        mst_list.loc[id,'Pur Unit']=new[prod][3]
        mst_list.loc[id,'Pur Qty']=new[prod][4]
        mst_list.loc[id,'Source']=new[prod][6]
        mst_list.loc[id,'Date']='2023-{}-{}'.format(mon,str(calendar.mdays[int(mon)]))
  
    else:
        if prod not in shown:
            shown[prod]=0
            mst_list.loc[id,'Supplier SKU']=new[prod][0][1]
            mst_list.loc[id,'Supplier Name']=new[prod][0][0].rstrip()
            mst_list.loc[id,'Price']=new[prod][0][2]
            mst_list.loc[id,'Similarity']=new[prod][0][5]
            mst_list.loc[id,'Pur Unit']=new[prod][0][3]
            mst_list.loc[id,'Pur Qty']=new[prod][0][4]
            mst_list.loc[id,'Source']=new[prod][0][6]
            mst_list.loc[id,'Date']='{}-{}-{}'.format(str(datetime.date.today().year),mon,str(calendar.mdays[int(mon)]))
        else:
            shown[prod] += 1
            mst_list.loc[id,'Supplier SKU']=new[prod][shown[prod]][1]
            mst_list.loc[id,'Supplier Name']=new[prod][shown[prod]][0].rstrip()
            mst_list.loc[id,'Price']=new[prod][shown[prod]][2]
            mst_list.loc[id,'Similarity']=new[prod][shown[prod]][5]
            mst_list.loc[id,'Pur Unit']=new[prod][shown[prod]][3]
            mst_list.loc[id,'Pur Qty']=new[prod][shown[prod]][4]
            mst_list.loc[id,'Source']=new[prod][shown[prod]][6]
            mst_list.loc[id,'Date']='{}-{}-{}'.format(str(datetime.date.today().year),mon,str(calendar.mdays[int(mon)]))
            
            
# Handle the data before export
mst=mst_list[['Supplier','Product','Supplier Name','SKU','Supplier SKU','Unit','Pur Unit','Bin Qty','Bin Size','Unit Qty','Pur Qty','Price','Bin Turn','Location','Similarity','Source','Alternative','Group','Note','Date','Growth R']]
mst=mst.sort_values(by=['Supplier','Product'])


# In[21]:


# Open terminal file and set the visualization format
#book=load_workbook(mst_path)
with pd.ExcelWriter(mst_path, mode='a',engine='openpyxl',if_sheet_exists="replace") as writer:
    mst.to_excel(writer,sheet_name=month,index=False)
    workbook = writer.book
    worksheet = writer.sheets[month]
    
# Currency format, text wrap format, text central, float format, row height, column width    
    worksheet.freeze_panes = 'A2'
    worksheet.column_dimensions['A'].width=13
    worksheet.column_dimensions['B'].width=51
    worksheet.column_dimensions['C'].width=51
    worksheet.column_dimensions['D'].width=13
    worksheet.column_dimensions['E'].width=13
    worksheet.column_dimensions['F'].width=9
    worksheet.column_dimensions['G'].width=9
    worksheet.column_dimensions['H'].width=9
    worksheet.column_dimensions['I'].width=9
    worksheet.column_dimensions['J'].width=9
    worksheet.column_dimensions['K'].width=9
    worksheet.column_dimensions['L'].width=13
    worksheet.column_dimensions['M'].width=11
    worksheet.column_dimensions['N'].width=11
    worksheet.column_dimensions['O'].width=11
    worksheet.column_dimensions['P'].width=11
    worksheet.column_dimensions['R'].width=11
    worksheet.column_dimensions['S'].width=51
    worksheet.column_dimensions['T'].width=11
    worksheet.column_dimensions['U'].width=11
    
    for l in enumerate(mst['Product']):
        cell=worksheet['U{}'.format(str(l[0]+2))]
        cell.number_format = '0.00%'
        try:
            lindex=last[last['Product']==l[1]].index.tolist()[0]+2
            cell.value='=({3}!I{2}+{3}!K{2}-{0}!I{1}-{0}!K{1})/({0}!I{1}+{0}!K{1})'.format(mst_sheet, lindex, str(l[0]+2),month)
        except IndexError:
            try:
                lindex=last[last['Product']==renew['name (required)'][renew[renew['new name']==l[1]].index.tolist()[0]]].index.tolist()[0]
                cell.value='=({3}!I{2}+{3}!K{2}-{0}!I{1}-{0}!K{1})/({3}!I{2}+{0}!K{1})'.format(mst_sheet, lindex, str(l[0]+2),month)
            except IndexError:
                cell.value='=0'
    
    align = Alignment(horizontal='center',vertical='bottom',wrap_text=True,shrink_to_fit=True,indent=0)
    for col in worksheet.columns:
        for c in col:
            c.alignment = align
    for cell in worksheet['B']:
        cell.alignment = Alignment(horizontal='left',vertical='bottom',text_rotation=0,wrap_text=True,shrink_to_fit=True,indent=0)
    for cell in worksheet['C']:
        cell.alignment = Alignment(horizontal='left',vertical='bottom',text_rotation=0,wrap_text=True,shrink_to_fit=True,indent=0)
    for cell in worksheet['L']:
        cell.number_format='"$"#,##0.00_);("$"#,##0.00)'
        cell.alignment = align
    for cell in enumerate(worksheet['M']):
        cell[1].number_format='#,##0.00'
        cell[1].value="=K{}/(H{}*I{}/2)".format(str(cell[0]+1),str(cell[0]+1),str(cell[0]+1))
        cell[1].alignment = align
    worksheet['M1'].value='Bin Turn'
    
    for cell in worksheet['1']:
        cell.fill = PatternFill(fill_type = 'solid',start_color='98FB98')
        cell.alignment=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,indent=0)
        
    worksheet['S2'].value='Check supplier name/SKU, Pur unit/QTY/price due to overwriting. And renew name & SKU'
    worksheet['S2'].font=Font(size=15,bold=True,color='9c0006')


# Once we pick the wrong match, from purchase history to monthly report, we can use <font color = blue>0 Similarity and None Supplier</font> to mark the wrong product that should be 0 purchase.         
# Also use <font color = blue>0 Similarity with Supplier</font> to mark the manually corrected product.           
# The following Code helps correct the wrong product. 
# # Products with alternative need to adjust the number manully. (same  name)

# In[22]:


with pd.ExcelWriter(mst_path, mode='a',engine='openpyxl',if_sheet_exists='replace') as writer:
    workbook = writer.book
    w4 = writer.sheets[month]
    w3 = writer.sheets[mst_sheet]

    # run the sheet to find the product where the similarity is 0 and no source, which means error
    for row in w4.iter_rows(min_row=1, min_col=15,max_col=16):
        s_s=[]
        for cell in row:         
            s_s.append(cell.value)
            if len(s_s)==2 and s_s[0]==0 and not s_s[1]:
                # find the thing that we are looking for 
                product=w4['B{}'.format(cell.row)].value
                count=0
                # open the previous report to pick the old b 
                for rows in w3.iter_rows(min_row=1, min_col=2,max_col=2):
                    for cells in rows:
                        if cells.value == product:
                            
                            # no alternative
                            if w4['Q{}'.format(cell.row)].value is None:
                                # name, sku, unit, qty, price

                                w4['C{}'.format(cell.row)].value = w3['C{}'.format(cells.row)].value
                                w4['E{}'.format(cell.row)].value = w3['E{}'.format(cells.row)].value
                                w4['G{}'.format(cell.row)].value = w3['G{}'.format(cells.row)].value
                                w4['K{}'.format(cell.row)].value = 0
                                w4['L{}'.format(cell.row)].value = w3['L{}'.format(cells.row)].value
                                count=1
                                break
                            # with alternatives    
                            else:
                                if w4['A{}'.format(cell.row)].value == w3['A{}'.format(cells.row)].value:
                                    if w4['Q{}'.format(cell.row)].value == w3['Q{}'.format(cells.row)].value:
                                        w4['C{}'.format(cell.row)].value = w3['C{}'.format(cells.row)].value
                                        w4['E{}'.format(cell.row)].value = w3['E{}'.format(cells.row)].value
                                        w4['G{}'.format(cell.row)].value = w3['G{}'.format(cells.row)].value
                                        w4['K{}'.format(cell.row)].value = 0
                                        w4['L{}'.format(cell.row)].value = w3['L{}'.format(cells.row)].value
                                        count=1
                                        break
                                elif w4['A{}'.format(cell.row)].value != w3['A{}'.format(cells.row)].value:
                                    continue
                            
                                
                                
                                
                # New product cannot be found in the previous sheet
                if count==0:
                    w4['C{}'.format(cell.row)].value = w4['B{}'.format(cell.row)].value
                    w4['E{}'.format(cell.row)].value = None
                    w4['G{}'.format(cell.row)].value = None
                    w4['K{}'.format(cell.row)].value = 0
                    w4['L{}'.format(cell.row)].value = 0
    # calculate growth rate again                
    for l in enumerate(mst['Product']):
        cell=w4['U{}'.format(str(l[0]+2))]
        try:
            lindex=last[last['Product']==l[1]].index.tolist()[0]+2
            cell.value='=({3}!I{2}+{3}!K{2}-{0}!I{1}-{0}!K{1})/({0}!I{1}+{0}!K{1})'.format(mst_sheet, lindex, str(l[0]+2),month)
            w4['R{}'.format(str(l[0]+2))].value='={0}!R{1}'.format(mst_sheet, lindex)
        except IndexError:
            try:
                lindex=last[last['Product']==renew['name (required)'][renew[renew['new name']==l[1]].index.tolist()[0]]].index.tolist()[0]
                cell.value='=({3}!I{2}+{3}!K{2}-{0}!I{1}-{0}!K{1})/({3}!I{2}+{0}!K{1})'.format(mst_sheet, lindex, str(l[0]+2),month)
                w4['R{}'.format(str(l[0]+2))].value='={0}!R{1}'.format(mst_sheet, lindex)
            except IndexError:
                cell.value='=0'
                w4['R{}'.format(str(l[0]+2))].value=np.nan


# In[ ]:




